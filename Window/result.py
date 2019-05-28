# -*- coding: utf-8 -*-
# -*- coding: utf-8 -*-
import time

from PyQt5.QtWidgets import QTableWidgetItem, QFileDialog

from Tool import xtredis
from Window.base import BaseWindow
from Tool.others import get_user_info, gender2str
from Tool.feedback import show_dialog
from UI.result import Ui_checked_result

import os
from openpyxl import Workbook

__author__ = "YingJoy"


class ResultWindow(BaseWindow):
    def __init__(self):
        super().__init__()

        self.ui = Ui_checked_result()
        self.ui.setupUi(self)
        self.init()
        [self.ui.result_table.setColumnWidth(i, 300) for i in range(self.ui.result_table.columnCount())]

        self.ui.export_btn.clicked.connect(self.export_result)

    def update(self):
        """更新数据"""
        # 先删除所有数据，这里不考虑性能
        self.ui.result_table.setRowCount(0)
        checked_uids = xtredis.lrange('checked_uid', 0, -1)
        checked_times = xtredis.lrange('checked_time', 0, -1)
        for index, uid in enumerate(checked_uids):
            if uid[:3] == 'sid':
                self.insert_row(
                    index,
                    [uid[3:], checked_times[index], '', '', '', '']
                )
            else:
                userinfo = get_user_info(uid)
                self.insert_row(
                    index,
                    [userinfo.get('sid'), checked_times[index], userinfo.get('name'),
                     gender2str(userinfo.get('gender')), userinfo.get('class_name'), userinfo.get('major')]
                )

        self.show()

    def insert_row(self, row_index, value):
        """插入一行数据"""
        self.ui.result_table.insertRow(row_index)
        self.ui.result_table.setRowHeight(row_index, 50)
        for col_index, item in enumerate(value):
            self.ui.result_table.setItem(row_index, col_index, QTableWidgetItem(item))

    def export_result(self):
        try:
            filepath = QFileDialog.getSaveFileName(self, "Save File", os.path.join(
                os.path.expanduser("~"),
                'Desktop/%s.xlsx' % str(int(time.time()))
            ), "All files (*.*);;")[0]

            workbook = Workbook()
            booksheet = workbook.active

            checked_uids = xtredis.lrange('checked_uid', 0, -1)
            checked_times = xtredis.lrange('checked_time', 0, -1)
            for row_index, uid in enumerate(checked_uids):
                if uid[:3] == 'sid':
                    for col_index, item in enumerate(
                            [uid[3:], checked_times[row_index], '', '', '', '']):
                        booksheet.cell(row_index + 1, col_index + 1).value = item
                else:
                    userinfo = get_user_info(uid)

                    for col_index, item in enumerate(
                            [userinfo.get('sid'), checked_times[row_index], userinfo.get('name'),
                             gender2str(userinfo.get('gender')), userinfo.get('class_name'), userinfo.get('major')]):
                        booksheet.cell(row_index + 1, col_index + 1).value = item

            workbook.save(filepath)
            show_dialog('Success', 'Export data success.')


        except Exception as e:
            show_dialog('Warn', 'Export fail.')
            print(e)

        self.close()
